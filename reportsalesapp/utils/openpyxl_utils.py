from tempfile import NamedTemporaryFile
import openpyxl
import json
from reportsalesapp.models import RealizationReport, RealizationReportDetail
from . import os_utils
from pathlib import Path
from django.core.files import File
import datetime


def open_excel_file(link: str, work_sheet=None):
    f = openpyxl.load_workbook(link)
    if work_sheet is None:
        sheet = f.active
    else:
        sheet = f[work_sheet]
    return f, sheet


def charfieldnull(value):
    result = '' if value is None else value
    return result


def datetimefielddata(value):
    data = datetime.date.fromisoformat(value)
    return data


def parsing_realization_report(company, file_link, set_report_number):
    file, sheet = open_excel_file(file_link)
    data = []
    cursor_row = 2
    cursor_cell_value = 0

    while cursor_cell_value is not None:
        report_number = int(sheet.cell(row=cursor_row, column=1).value)
        if realizationreport_id not in set_report_number:
            report = RealizationReport(
                number=report_number,
                date_from=sheet.cell(row=cursor_row, column=3).value,
                date_to=sheet.cell(row=cursor_row, column=4).value,
                create_dt=sheet.cell(row=cursor_row, column=5).value,
                sales=sheet.cell(row=cursor_row, column=6).value,
                transfer_for_goods=sheet.cell(row=cursor_row, column=8).value,
                logistic=sheet.cell(row=cursor_row, column=9).value,
                penalty_logistic=sheet.cell(row=cursor_row, column=10).value,
                penalty_other=sheet.cell(row=cursor_row, column=11).value,
                penalty=sheet.cell(row=cursor_row, column=12).value,
                additional_payment=sheet.cell(row=cursor_row, column=13).value,
                storage=sheet.cell(row=cursor_row, column=14).value,
                acceptance=sheet.cell(row=cursor_row, column=15).value,
                deduction=sheet.cell(row=cursor_row, column=16).value,
                total=sheet.cell(row=cursor_row, column=17).value,
                currency=sheet.cell(row=cursor_row, column=18).value,
                company=company
            )

            data.append(report)
        cursor_row += 1
        cursor_cell_value = sheet.cell(row=cursor_row, column=1).value

    return data


def parsing_realization_report_detail(file_link, realization_report):
    file, sheet = open_excel_file(file_link)
    data = []
    cursor_row = 2
    cursor_cell_value = 0
    print('start parsing' + ' ' + str({realization_report.number}))

    while cursor_cell_value is not None:
        if (sheet.cell(row=cursor_row, column=11).value not in {
            'Возмещение издержек по перевозке/по складским операциям с товаром',
            'Возмещение издержек по перевозке',
        }):
            line = RealizationReportDetail(
                realizationReport=realization_report,
                gi_id=sheet.cell(row=cursor_row, column=2).value,
                subject_name=charfieldnull(sheet.cell(row=cursor_row, column=3).value),
                nm_id=sheet.cell(row=cursor_row, column=4).value,
                brand_name=charfieldnull(sheet.cell(row=cursor_row, column=5).value),
                sa_name=charfieldnull(sheet.cell(row=cursor_row, column=6).value),
                ts_name=charfieldnull(sheet.cell(row=cursor_row, column=8).value),
                barcode=charfieldnull(sheet.cell(row=cursor_row, column=9).value),
                doc_type_name=charfieldnull(sheet.cell(row=cursor_row, column=10).value),
                quantity=sheet.cell(row=cursor_row, column=14).value,
                retail_price=sheet.cell(row=cursor_row, column=15).value,
                retail_amount=sheet.cell(row=cursor_row, column=16).value,
                office_name=charfieldnull(sheet.cell(row=cursor_row, column=45).value),
                supplier_oper_name=charfieldnull(sheet.cell(row=cursor_row, column=11).value),
                order_dt=datetimefielddata(sheet.cell(row=cursor_row, column=12).value),
                sale_dt=datetimefielddata(sheet.cell(row=cursor_row, column=13).value),
                shk_id=sheet.cell(row=cursor_row, column=50).value,
                delivery_amount=sheet.cell(row=cursor_row, column=33).value,
                return_amount=sheet.cell(row=cursor_row, column=34).value,
                delivery_rub=sheet.cell(row=cursor_row, column=35).value,
                gi_box_type_name=charfieldnull(sheet.cell(row=cursor_row, column=47).value),
                ppvz_for_pay=sheet.cell(row=cursor_row, column=32).value,
                site_country=charfieldnull(sheet.cell(row=cursor_row, column=46).value),
                penalty=sheet.cell(row=cursor_row, column=36).value,
                additional_payment=sheet.cell(row=cursor_row, column=37).value,
                storage_fee=sheet.cell(row=cursor_row, column=55).value,
                deduction=sheet.cell(row=cursor_row, column=56).value,
                acceptance=sheet.cell(row=cursor_row, column=57).value,
            )
            data.append(line)

        cursor_row += 1
        cursor_cell_value = sheet.cell(row=cursor_row, column=1).value
        # print(cursor_row)
    print('end parsing')

    return data





