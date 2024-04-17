import openpyxl
import re
import json
from goodseller.settings import GENERAL_SALES_REPORT
from .openpyxl_utils import open_excel_file


def map_str_to_int(s):
    if isinstance(s, str):
        if s.startswith("="):
            return sum(int(x) for x in re.findall('[-+]?\d+', s[1::]))
    elif isinstance(s, int):
        return s
    else:
        return 0


def input_cell(sheet_report, column, row, data_old, data_new):
    if data_old is not None:
        if data_old.startswith("="):
            sheet_report[f"{column}{row}"] = data_old + '+' + data_new
        else:
            sheet_report[f"{column}{row}"] = "=" + data_old + '+' + data_new
    else:
        sheet_report(row=row, column=column).value = data_new


def control_report(dict_result):
    for row in dict_result:
        if row['count'] != 0 or row['sales'] != 0 \
                or row['commission'] != 0 or row['logistic'] != 0:
            print(row)
    print("END CONTROL")


def control_data(data):
    for row in data:
        if row['count'] != 0 or row['sales'] != 0 \
                or row['commission'] != 0 or row['logistic'] != 0:
            return False
    return True


def input_data_general_report(data, firstiterations=True):
    file, sheet = open_excel_file(GENERAL_SALES_REPORT, 'Продажи_WB')

    def examination_articles(art1, art2):
        if isinstance(art1, int):
            return art1 == art2
        elif isinstance(art1, str):
            return art1.rfind(str(art2)) != -1
        else:
            return False

    def examination_availability_wb(availability_wb, firstiterations):
        return availability_wb > 0 if firstiterations else True

    for item in data:
        # print(item['nm_id'])
        cursor_row = 6
        cursor_cell_value = 1

        while cursor_cell_value is not None:
            article = sheet.cell(row=cursor_row, column=4).value
            try:
                availability_wb = map_str_to_int(sheet.cell(row=cursor_row, column=12).value) \
                                  - map_str_to_int(sheet.cell(row=cursor_row, column=14).value)
            except TypeError:
                print(f"По строке {cursor_row} возникла ошибка {item['nm_id']}")

            count_cell = sheet.cell(row=cursor_row, column=14).value
            sales_cell = sheet.cell(row=cursor_row, column=15).value
            commission_cell = sheet.cell(row=cursor_row, column=17).value
            logistic_cell = sheet.cell(row=cursor_row, column=18).value

            if examination_articles(article, item['nm_id']) and \
                    examination_availability_wb(availability_wb, firstiterations):
                if item['count'] > availability_wb > 0:
                    temp_count = availability_wb
                    temp_sales = round(item['sales'] / item['count'] * temp_count, 2)
                    temp_commission = round(item['commission'] / item['count'] * temp_count, 2)
                    temp_logistic = round(item['logistic'] / item['count'] * temp_count, 2)
                    input_cell(sheet, "N", cursor_row, str(count_cell), str(temp_count))
                    input_cell(sheet, "O", cursor_row, str(sales_cell), str(temp_sales))
                    input_cell(sheet, "Q", cursor_row, str(commission_cell), str(temp_commission))
                    input_cell(sheet, "R", cursor_row, str(logistic_cell), str(temp_logistic))
                    item['count'] -= temp_count
                    item['sales'] -= temp_sales
                    item['commission'] -= temp_commission
                    item['logistic'] -= temp_logistic

                else:
                    input_cell(sheet, "N", cursor_row, str(count_cell), str(item['count']))
                    input_cell(sheet, "O", cursor_row, str(sales_cell), str(item['sales']))
                    input_cell(sheet, "Q", cursor_row, str(commission_cell), str(item['commission']))
                    input_cell(sheet, "R", cursor_row, str(logistic_cell), str(item['logistic']))
                    item['count'] = 0
                    item['sales'] = 0
                    item['commission'] = 0
                    item['logistic'] = 0
            cursor_row += 1
            cursor_cell_value = sheet.cell(row=cursor_row, column=1).value

    control_report(data)
    file.save(GENERAL_SALES_REPORT)
